import openpyxl
import os

def save_to_excel(roll_number, name, age, grade):
    file_path = "student_data.xlsx"
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        # Check if the sheet with title 'Student Data' already exists
        if 'Student Data' in workbook.sheetnames:
            sheet = workbook['Student Data']
        else:
            sheet = workbook.create_sheet(title="Student Data", index=0)
            # Set column headers if the sheet is empty
            sheet.cell(1, 1, "Roll Number")
            sheet.cell(1, 2, "Name")
            sheet.cell(1, 3, "Age")
            sheet.cell(1, 4, "Grade")
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Student Data"
        # Set column headers for the new sheet
        sheet.cell(1, 1, "Roll Number")
        sheet.cell(1, 2, "Name")
        sheet.cell(1, 3, "Age")
        sheet.cell(1, 4, "Grade")

    # Find the next empty row to add data
    row_number = sheet.max_row + 1
    sheet.cell(row_number, 1, roll_number)
    sheet.cell(row_number, 2, name)
    sheet.cell(row_number, 3, age)
    sheet.cell(row_number, 4, grade)

    # Save the data to the Excel file
    workbook.save(file_path)

if __name__ == "__main__":
    try:
        roll_number = int(input("Enter Roll Number: "))
        name = input("Enter Name: ")

        while True:
            try:
                age = int(input("Enter Age: "))
                break  # Break the loop if age input is valid
            except ValueError:
                print("Invalid age. Please enter a valid integer.")

        grade = input("Enter Grade: ")

        save_to_excel(roll_number, name, age, grade)

        print("Data saved successfully to 'student_data.xlsx'.")
    except ValueError:
        print("Invalid input. Please enter valid data.")